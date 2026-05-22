"""
services/report_writer.py — [섹션 6] Excel 컬러 리포트 생성
app.py의 _build_view_sheet, build_report를 그대로 분리. 로직 수정 없음.
"""
from __future__ import annotations
import logging
import re
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from backend.services.utils import _title_contains_view, _축척_텍스트_정리

logger = logging.getLogger("AutoDWG")


def _build_view_sheet(ws, view_df: pd.DataFrame) -> None:
    """뷰심볼 검토 시트를 작성한다."""
    빨간색 = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    주황색 = PatternFill(start_color="FFFFD699", end_color="FFFFD699", fill_type="solid")
    헤더색 = PatternFill(start_color="FFD6E4F7", end_color="FFD6E4F7", fill_type="solid")

    cols = {
        "파일명":        "파일명",
        "도면명(DWG)":   "도곽 도면명",
        "축척_A1(DWG)":  "도곽 A1축척",
        "축척_A3(DWG)":  "도곽 A3축척",
        "뷰_도면명":     "뷰 도면명",
        "뷰_A1축척":     "뷰 A1축척",
        "뷰_A3축척":     "뷰 A3축척",
        "도면명_포함":   "도면명 포함",
        "축척_일치":     "축척 일치",
        "상태":          "상태",
    }
    df = view_df.copy()

    def _chk_title(row):
        return "O" if _title_contains_view(str(row.get("도면명(DWG)", "")), str(row.get("뷰_도면명", ""))) else "X"

    def _chk_scale(row):
        v_a1 = re.sub(r"[,\s]", "", str(row.get("뷰_A1축척", ""))).upper()
        v_a3 = re.sub(r"[,\s]", "", str(row.get("뷰_A3축척", ""))).upper()
        d_a1 = re.sub(r"[,\s]", "", str(row.get("축척_A1(DWG)", ""))).upper()
        d_a3 = re.sub(r"[,\s]", "", str(row.get("축척_A3(DWG)", ""))).upper()
        if not v_a1 and not v_a3:
            return "?"
        results = []
        if v_a1 and v_a1 not in ("X", "NONE"):
            results.append(v_a1 == d_a1)
        if v_a3 and v_a3 not in ("X", "NONE"):
            results.append(v_a3 == d_a3)
        if not results:
            return "?"
        return "O" if all(results) else "X"

    df["도면명_포함"] = df.apply(_chk_title, axis=1)
    df["축척_일치"]  = df.apply(_chk_scale, axis=1)

    def _view_status(r):
        title_ok = r["도면명_포함"] == "O"
        scale_ok = r["축척_일치"]  == "O"
        if title_ok and scale_ok:
            return "일치"
        parts = []
        if not title_ok:
            parts.append("도면명")
        if not scale_ok:
            parts.append("축척")
        return "/".join(parts) + " 불일치"

    df["상태"] = df.apply(_view_status, axis=1)

    # 같은 파일 내 뷰 도면명이 중복이면 오타 가능성 → "중복"으로 덮어쓰기
    dup_mask = df.duplicated(subset=["파일명", "뷰_도면명"], keep=False)
    df.loc[dup_mask, "상태"] = "중복"

    col_keys = list(cols.keys())
    for j, (key, label) in enumerate(cols.items(), 1):
        c = ws.cell(1, j, label)
        c.fill = 헤더색
    for i, row in enumerate(df[col_keys].fillna("").itertuples(index=False), 2):
        for j, val in enumerate(row, 1):
            ws.cell(i, j, str(val))
        status = ws.cell(i, col_keys.index("상태") + 1).value
        if status == "중복":
            for j in range(1, len(col_keys) + 1):
                ws.cell(i, j).fill = 주황색
        elif status != "일치":
            for j in range(1, len(col_keys) + 1):
                ws.cell(i, j).fill = 빨간색


def build_report(
    list_df: pd.DataFrame,
    dwg_df: pd.DataFrame,
    out_path: str,
    view_df: Optional[pd.DataFrame] = None
) -> None:
    """도면목록표 검토 결과를 Excel 파일로 저장한다.

    시트1: 목록표 검토 (LIST ↔ DWG 대조)
    시트2: 뷰심볼 검토 (view_df가 있을 경우)
    """
    if list_df.empty and dwg_df.empty:
        logger.warning("[알림] 추출된 데이터가 없어 엑셀 리포트를 생성하지 않습니다.")
        return

    lst = list_df.copy()
    dwg = dwg_df.copy()

    # 필수 컬럼 보장
    if "도면번호(LIST)"  not in lst.columns: lst["도면번호(LIST)"]  = ""
    if "도면번호(DWG)"   not in dwg.columns: dwg["도면번호(DWG)"]   = ""
    if "구분_LIST(그룹)" not in lst.columns: lst["구분_LIST(그룹)"] = ""
    if "구분_DWG(그룹)"  not in dwg.columns: dwg["구분_DWG(그룹)"]  = ""

    # 도면번호 정규화 키를 기준으로 outer merge
    lst["KEY"] = lst["도면번호(LIST)"].astype(str).str.upper().str.replace(r"[\s\-_]", "", regex=True)
    dwg["KEY"] = dwg["도면번호(DWG)"].astype(str).str.upper().str.replace(r"[\s\-_]", "", regex=True)
    결과 = pd.merge(lst, dwg, on="KEY", how="outer", indicator=True)
    결과["상태"] = 결과["_merge"].map({"both": "일치", "left_only": "DWG 누락", "right_only": "목록표 누락"})

    # 그룹 불일치 행 인덱스 수집 (Excel 행 번호 기준: 헤더=1, 데이터=2~)
    group_mismatch_indices = set()
    for i in range(len(결과)):
        l_g = str(결과.at[i, "구분_LIST(그룹)"]).strip()
        d_g = str(결과.at[i, "구분_DWG(그룹)"]).strip()
        if l_g == "nan": l_g = ""
        if d_g == "nan": d_g = ""
        if l_g and d_g and l_g != d_g:
            group_mismatch_indices.add(i + 2)

    # 연속으로 같은 그룹이 반복되면 빈칸으로 (시각적 그룹핑)
    prev_group = ""
    group_col_idx = 결과.columns.get_loc("구분_LIST(그룹)")
    for i in range(len(결과)):
        curr_group = str(결과.iat[i, group_col_idx]).strip()
        if curr_group == "nan" or not curr_group:
            prev_group = ""
            결과.iat[i, group_col_idx] = ""
            continue
        if curr_group == prev_group:
            결과.iat[i, group_col_idx] = ""
        else:
            prev_group = curr_group

    prev_dwg_group = ""
    dwg_group_col_idx = 결과.columns.get_loc("구분_DWG(그룹)")
    for i in range(len(결과)):
        curr_group = str(결과.iat[i, dwg_group_col_idx]).strip()
        if curr_group == "nan" or not curr_group:
            prev_dwg_group = ""
            결과.iat[i, dwg_group_col_idx] = ""
            continue
        if curr_group == prev_dwg_group:
            결과.iat[i, dwg_group_col_idx] = ""
        else:
            prev_dwg_group = curr_group

    cols = [
        "도면번호(LIST)", "구분_LIST(그룹)", "도면명(LIST)", "축척_A1(LIST)", "축척_A3(LIST)",
        "도면번호(DWG)",  "구분_DWG(그룹)",  "도면명(DWG)",  "축척_A1(DWG)",  "축척_A3(DWG)",
        "파일명", "상태"
    ]
    for c in cols:
        if c not in 결과.columns:
            결과[c] = ""

    결과[cols].fillna("X").to_excel(out_path, index=False)

    # openpyxl로 컬러 적용
    wb = load_workbook(out_path)
    ws = wb.active
    ws.title = "목록표 검토"
    빨간색 = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    h = {cell.value: cell.column for cell in ws[1] if cell.value}

    for row in range(2, ws.max_row + 1):
        status = ws.cell(row, h["상태"]).value
        # DWG 누락 / 목록표 누락 → 행 전체 빨간색
        if status in ("DWG 누락", "목록표 누락"):
            for c in range(1, len(cols) + 1):
                ws.cell(row, c).fill = 빨간색
            continue
        # 항목별 불일치 검사
        issues = []
        if row in group_mismatch_indices:
            issues.append("그룹")
            if h.get("구분_LIST(그룹)"): ws.cell(row, h["구분_LIST(그룹)"]).fill = 빨간색
            if h.get("구분_DWG(그룹)"): ws.cell(row, h["구분_DWG(그룹)"]).fill  = 빨간색
        val_list = re.sub(r"[\s\-_]", "", str(ws.cell(row, h["도면번호(LIST)"]).value).upper())
        val_dwg  = re.sub(r"[\s\-_]", "", str(ws.cell(row, h["도면번호(DWG)"]).value).upper())
        if val_list != val_dwg:
            issues.append("도면번호")
            ws.cell(row, h["도면번호(LIST)"]).fill = 빨간색
            ws.cell(row, h["도면번호(DWG)"]).fill  = 빨간색
        name_list = str(ws.cell(row, h["도면명(LIST)"]).value).replace(" ", "")
        name_dwg  = str(ws.cell(row, h["도면명(DWG)"]).value).replace(" ", "")
        if name_list != name_dwg:
            issues.append("도면명")
            ws.cell(row, h["도면명(LIST)"]).fill = 빨간색
            ws.cell(row, h["도면명(DWG)"]).fill  = 빨간색
        scale_bad = False
        for s in ["A1", "A3"]:
            p_v = str(ws.cell(row, h[f"축척_{s}(LIST)"]).value).replace(" ", "")
            d_v = str(ws.cell(row, h[f"축척_{s}(DWG)"]).value).replace(" ", "")
            if p_v != d_v:
                scale_bad = True
                ws.cell(row, h[f"축척_{s}(LIST)"]).fill = 빨간색
                ws.cell(row, h[f"축척_{s}(DWG)"]).fill  = 빨간색
        if scale_bad:
            issues.append("축척")
        if issues:
            ws.cell(row, h["상태"]).value = "/".join(issues) + " 불일치"
            ws.cell(row, h["상태"]).fill  = 빨간색

    # 시트2: 뷰심볼 검토
    if view_df is not None and not view_df.empty:
        ws_view = wb.create_sheet("뷰심볼 검토")
        _build_view_sheet(ws_view, view_df)
        logger.info("[XLSX] 뷰심볼 검토 시트 추가: %d건", len(view_df))

    wb.save(out_path)
    logger.info("[XLSX] 리포트 저장 완료: %s", out_path)
